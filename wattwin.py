# wattwin.py
import os
import base64
import json
import requests
from openpyxl import Workbook
from io import BytesIO, StringIO

def get_order_id(instance_id: str, api_key: str):
    url = "https://public.api.wattwin.com/v1/ECommerceOrders/search"
    payload = {
        "query": {
            "term": {"instanceId": instance_id}
        }
    }
    resp = requests.post(url, headers={"Content-Type": "application/json", "x-api-key": api_key}, json=payload)
    resp.raise_for_status()
    hits = resp.json().get("hits", {}).get("hits", [])
    if not hits:
        raise ValueError(f"No se encontró order para instanceId {instance_id}")
    return hits[0]["_source"]["id"]

def process_wattwin_order(instance_id: str, nombre: str, fecha: str, ref: str):
    WATTWIN_API_KEY = os.environ["WATTWIN_API_KEY"]
    order_id = get_order_id(instance_id, WATTWIN_API_KEY)
    
    log_stream = StringIO()
    def log(msg):
        log_stream.write(msg + "\n")

    WATTWIN_API_KEY = os.environ["WATTWIN_API_KEY"]
    GITHUB_TOKEN = os.environ["GITHUB_TOKEN"]
    GITHUB_REPO = "AlbertoMayoralGreenSunrise/Render"
    GITHUB_BRANCH = "main"

    github_api_url_excel = f"https://api.github.com/repos/{GITHUB_REPO}/contents/TEST_DEFINITIVO.xlsx"
    headers = {"Authorization": f"Bearer {GITHUB_TOKEN}"}

    # --- Llamar a Wattwin ---
    try:
        resp = requests.get(
            "https://public.api.wattwin.com/v1/ECommerceOrderLines",
            headers={"accept": "application/json", "x-api-key": WATTWIN_API_KEY},
            params={"filter": f'{{"where":{{"orderId":"{order_id}"}}}}'}
        )
        resp.raise_for_status()
        products_lines = resp.json()
    except Exception as e:
        products_lines = []

    # --- Crear Excel nuevo ---
    columns = [
        "Numero", "Nombre",
        "Estructura", "Paneles", "Unidades Estructura/Paneles",
        "Optimizador", "Unidades Optimizador",
        "Inversor", "Unidades Inversor",
        "Baterías", "Unidades Baterías",
        "Cargador VE", "Pajareras",
        "Fecha de venta"
    ]

    from openpyxl import load_workbook
    
    # --- Cargar Excel desde GitHub si existe ---
    wb = None
    get_resp = requests.get(github_api_url_excel, headers=headers)
    if get_resp.status_code == 200:
        content = base64.b64decode(get_resp.json()["content"])
        wb = load_workbook(filename=BytesIO(content))
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"
        ws.append(columns)
    
    ws = wb.active

    # --- Mapeo categoryId → columna ---
    category_to_column = {
        "6328b2a5efa9419a5938b92d": 2,  # Estructura
        "6328b2a5efa9419a5938b91c": 3,  # Paneles
        "6790e34a0a5301a6d0b6e7f8": 5,  # Optimizador
        "6328b2a5efa9419a5938b921": 7,  # Inversor
        "6328b2a5efa9419a5938b927": 9,  # Batería
        "678e12f76d2390929fd91374": 11, # Cargador VE
        "6328b2a5efa9419a5938b92f": 12  # Pajareras
    }

    category_to_units_column = {
        "6328b2a5efa9419a5938b92d": 4,  # Unidades Estructura
        "6328b2a5efa9419a5938b91c": 4,  # Unidades Paneles
        "6790e34a0a5301a6d0b6e7f8": 6,  # Unidades Optimizador
        "6328b2a5efa9419a5938b921": 8,  # Unidades Inversor
        "6328b2a5efa9419a5938b927": 10, # Unidades Batería
        # Cargador VE y Pajareras no tienen columna de unidades, se puede omitir
    }


    # --- Crear fila del pedido ---
    pedido_row = [""] * len(columns)
    pedido_row[0] = ref        # Columna "Numero"
    pedido_row[1] = nombre     # Columna "Nombre"
    pedido_row[13] = fecha     # Columna "Fecha de venta"


    for idx, line in enumerate(products_lines, start=1):
        product_name = line.get("name", "")
        count = line.get("count", 0)
        product_id = line.get("productId")
        
        category_id = ""
        if product_id:
            try:
                product_resp = requests.get(
                    f"https://public.api.wattwin.com/v1/Products/{product_id}",
                    headers={"accept": "application/json", "x-api-key": WATTWIN_API_KEY}
                )
                product_resp.raise_for_status()
                product_data = product_resp.json()
                category_id = product_data.get("categoryId", "")
            except Exception as e:
                log(f"[ERROR] No se pudo obtener categoría del producto {product_id}: {e}")


        if category_id in category_to_column:
            col_idx = category_to_column[category_id]
        
            # --- Caso especial: Pajareras ---
            if category_id == "6328b2a5efa9419a5938b92f":  # Pajareras
                pedido_row[col_idx] = "Sí"
            else:
                units_col_idx = category_to_units_column.get(category_id)
                if pedido_row[col_idx]:
                    pedido_row[col_idx] += f", {product_name}"
                    if units_col_idx:
                        pedido_row[units_col_idx] += f" + {count}"
                else:
                    pedido_row[col_idx] = product_name
                    if units_col_idx:
                        pedido_row[units_col_idx] = str(count)


                
    # --- Insertar en la primera fila vacía ---
    next_row = ws.max_row + 1
    for col_idx, value in enumerate(pedido_row, start=1):
        ws.cell(row=next_row, column=col_idx, value=value)
    log(f"[LOG] Fila agregada al Excel en fila {next_row}: {pedido_row}")

    # --- Guardar Excel en GitHub (con sobreescritura) ---
    output = BytesIO()
    wb.save(output)
    content_excel = base64.b64encode(output.getvalue()).decode()
    
    # Obtener sha si el archivo existe
    sha = None
    get_resp = requests.get(github_api_url_excel, headers=headers)
    if get_resp.status_code == 200:
        sha = get_resp.json()["sha"]
    
    data_excel = {
        "message": f"Crear/Actualizar Excel para pedido {order_id}",
        "content": content_excel,
        "branch": GITHUB_BRANCH,
    }
    if sha:
        data_excel["sha"] = sha  # necesario para sobreescribir
    
    try:
        put_resp = requests.put(github_api_url_excel, headers=headers, data=json.dumps(data_excel))
        put_resp.raise_for_status()
    except Exception as e:
        log(f"[ERROR] GitHub PUT falló: {e}")

    
